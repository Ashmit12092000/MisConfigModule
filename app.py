import os
import bcrypt
from datetime import datetime, date
from flask import Flask, render_template, request, redirect, url_for, session, flash
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from sqlalchemy import event

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SESSION_SECRET', 'dev-secret-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///mis_config.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)

class Role(db.Model):
    __tablename__ = 'roles'
    RoleID = db.Column(db.Integer, primary_key=True)
    RoleName = db.Column(db.String(50), unique=True, nullable=False)
    users = db.relationship('User', backref='role', lazy=True)

class Company(db.Model):
    __tablename__ = 'companies'
    CompanyID = db.Column(db.Integer, primary_key=True)
    CompanyName = db.Column(db.String(100), unique=True, nullable=False)
    ActiveFlag = db.Column(db.Boolean, default=True)

class Department(db.Model):
    __tablename__ = 'departments'
    DeptID = db.Column(db.Integer, primary_key=True)
    DeptName = db.Column(db.String(100), unique=True, nullable=False)
    ActiveFlag = db.Column(db.Boolean, default=True)
    users = db.relationship('User', backref='department', lazy=True)
    uploads = db.relationship('MISUpload', backref='department', lazy=True)
    templates = db.relationship('Template', backref='department', lazy=True)

class FinancialYear(db.Model):
    __tablename__ = 'financial_years'
    FYID = db.Column(db.Integer, primary_key=True)
    FYName = db.Column(db.String(50), unique=True, nullable=False)
    StartDate = db.Column(db.Date, nullable=False)
    EndDate = db.Column(db.Date, nullable=False)
    ActiveFlag = db.Column(db.Boolean, default=False)
    uploads = db.relationship('MISUpload', backref='financial_year', lazy=True)

class User(db.Model):
    __tablename__ = 'users'
    UserID = db.Column(db.Integer, primary_key=True)
    Username = db.Column(db.String(50), unique=True, nullable=False)
    PasswordHash = db.Column(db.String(255), nullable=False)
    Email = db.Column(db.String(100), unique=True, nullable=False)
    DepartmentID = db.Column(db.Integer, db.ForeignKey('departments.DeptID'), nullable=False)
    RoleID = db.Column(db.Integer, db.ForeignKey('roles.RoleID'), nullable=False)
    IsActive = db.Column(db.Boolean, default=True)
    uploads = db.relationship('MISUpload', backref='uploader', lazy=True)

class MISUpload(db.Model):
    __tablename__ = 'mis_uploads'
    UploadID = db.Column(db.Integer, primary_key=True)
    DepartmentID = db.Column(db.Integer, db.ForeignKey('departments.DeptID'), nullable=False)
    MonthID = db.Column(db.Integer, nullable=False)
    FYID = db.Column(db.Integer, db.ForeignKey('financial_years.FYID'), nullable=False)
    UploadedBy = db.Column(db.Integer, db.ForeignKey('users.UserID'), nullable=False)
    UploadDate = db.Column(db.DateTime, default=datetime.utcnow)
    FilePath = db.Column(db.String(255), nullable=False)
    Status = db.Column(db.String(50), default='Pending')

class Template(db.Model):
    __tablename__ = 'templates'
    TemplateID = db.Column(db.Integer, primary_key=True)
    DepartmentID = db.Column(db.Integer, db.ForeignKey('departments.DeptID'), nullable=False)
    FilePath = db.Column(db.String(255), nullable=False)
    UploadDate = db.Column(db.DateTime, default=datetime.utcnow)

def hash_password(password):
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password, hashed):
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        user = User.query.get(session['user_id'])
        if not user or not user.IsActive:
            session.clear()
            flash('Your session has expired. Please log in again.', 'error')
            return redirect(url_for('login'))
        if user.role.RoleName != 'Admin':
            flash('Access denied. Admin privileges required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def hod_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        user = User.query.get(session['user_id'])
        if not user or not user.IsActive:
            session.clear()
            flash('Your session has expired. Please log in again.', 'error')
            return redirect(url_for('login'))
        if user.role.RoleName not in ['Admin', 'HOD']:
            flash('Access denied. HOD privileges required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def check_upload_window():
    today = date.today()
    if 1 <= today.day <= 10:
        return True, "Upload window is open (1st-10th of the month)."
    elif today.day > 10:
        return False, "Upload window closed for the current month. Window opens on the 1st of next month."
    else:
        return False, "Upload window opens on the 1st of the month."

def validate_excel_file(file_path):
    """Validate Excel file structure and content"""
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file_path)
        
        if len(workbook.sheetnames) == 0:
            return False, "Excel file has no sheets."
        
        sheet = workbook.active
        if sheet.max_row < 2:
            return False, "Excel file appears to be empty (no data rows)."
        
        if sheet.max_column < 1:
            return False, "Excel file has no columns."
        
        return True, "File validation successful."
    except Exception as e:
        return False, f"File validation error: {str(e)}"

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(Username=username).first()
        
        if user and user.IsActive and verify_password(password, user.PasswordHash):
            session['user_id'] = user.UserID
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid credentials or inactive account.', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully.', 'success')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    active_fy = FinancialYear.query.filter_by(ActiveFlag=True).first()
    
    # Get recent uploads based on role
    if user.role.RoleName == 'Admin':
        recent_uploads = MISUpload.query.order_by(MISUpload.UploadDate.desc()).limit(5).all()
    elif user.role.RoleName == 'HOD':
        recent_uploads = MISUpload.query.filter_by(DepartmentID=user.DepartmentID).order_by(MISUpload.UploadDate.desc()).limit(5).all()
    else:
        recent_uploads = MISUpload.query.filter_by(DepartmentID=user.DepartmentID).order_by(MISUpload.UploadDate.desc()).limit(5).all()
    
    stats = {
        'total_users': User.query.count(),
        'total_depts': Department.query.count(),
        'active_fy': active_fy.FYName if active_fy else 'None',
        'total_uploads': MISUpload.query.count(),
        'recent_uploads': recent_uploads
    }
    
    return render_template('dashboard.html', current_user=user, stats=stats)


@app.route('/reports')
@login_required
def reports():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Filter uploads based on role
    if user.role.RoleName == 'Admin':
        uploads = MISUpload.query.order_by(MISUpload.UploadDate.desc()).all()
    else:
        uploads = MISUpload.query.filter_by(DepartmentID=user.DepartmentID).order_by(MISUpload.UploadDate.desc()).all()
    
    departments = Department.query.filter_by(ActiveFlag=True).all()
    financial_years = FinancialYear.query.all()
    
    return render_template('reports.html', 
                         current_user=user,
                         uploads=uploads,
                         departments=departments,
                         financial_years=financial_years)

@app.route('/approval-queue')
@hod_required
def approval_queue():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Get pending uploads based on role
    if user.role.RoleName == 'Admin':
        pending_uploads = MISUpload.query.filter(MISUpload.Status.in_(['Pending', 'Validated'])).order_by(MISUpload.UploadDate.desc()).all()
    else:
        pending_uploads = MISUpload.query.filter_by(DepartmentID=user.DepartmentID).filter(MISUpload.Status.in_(['Pending', 'Validated'])).order_by(MISUpload.UploadDate.desc()).all()
    
    return render_template('approval_queue.html', 
                         current_user=user,
                         pending_uploads=pending_uploads)

@app.route('/approve-upload/<int:upload_id>', methods=['POST'])
@hod_required
def approve_upload(upload_id):
    user = User.query.get(session['user_id'])
    upload = MISUpload.query.get_or_404(upload_id)
    
    # Check permissions
    if user.role.RoleName == 'HOD' and upload.DepartmentID != user.DepartmentID:
        flash('You can only approve uploads from your department.', 'error')
        return redirect(url_for('approval_queue'))
    
    upload.Status = 'Approved'
    db.session.commit()
    flash(f'Upload approved successfully!', 'success')
    return redirect(url_for('approval_queue'))

@app.route('/reject-upload/<int:upload_id>', methods=['POST'])
@hod_required
def reject_upload(upload_id):
    user = User.query.get(session['user_id'])
    upload = MISUpload.query.get_or_404(upload_id)
    
    # Check permissions
    if user.role.RoleName == 'HOD' and upload.DepartmentID != user.DepartmentID:
        flash('You can only reject uploads from your department.', 'error')
        return redirect(url_for('approval_queue'))
    
    upload.Status = 'Rejected'
    db.session.commit()
    flash(f'Upload rejected.', 'warning')
    return redirect(url_for('approval_queue'))

@app.route('/download-upload/<int:upload_id>')
@login_required
def download_upload(upload_id):
    user = User.query.get(session['user_id'])
    upload = MISUpload.query.get_or_404(upload_id)
    
    # Check permissions - Admin and HOD can download all, User can only download from their department
    if user.role.RoleName not in ['Admin', 'HOD'] and upload.DepartmentID != user.DepartmentID:
        flash('Access denied.', 'error')
        return redirect(url_for('reports'))
    
    from flask import send_file
    try:
        filename = upload.FilePath.split('/')[-1]
        return send_file(upload.FilePath, as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f'Error downloading file: {str(e)}', 'error')
        return redirect(url_for('reports'))

@app.route('/delete-upload/<int:upload_id>', methods=['POST'])
@hod_required
def delete_upload(upload_id):
    user = User.query.get(session['user_id'])
    upload = MISUpload.query.get_or_404(upload_id)
    
    # HOD can only delete from their department
    if user.role.RoleName == 'HOD' and upload.DepartmentID != user.DepartmentID:
        flash('You can only delete uploads from your department.', 'error')
        return redirect(url_for('reports'))
    
    # Delete the file from filesystem
    try:
        if os.path.exists(upload.FilePath):
            os.remove(upload.FilePath)
    except Exception as e:
        flash(f'Warning: File deletion error: {str(e)}', 'warning')
    
    db.session.delete(upload)
    db.session.commit()
    flash('Upload deleted successfully!', 'success')
    return redirect(url_for('reports'))

@app.route('/config-master')
@admin_required
def config_master():
    user = User.query.get(session['user_id'])
    companies = Company.query.all()
    departments = Department.query.all()
    financial_years = FinancialYear.query.all()
    
    return render_template('config_master.html', 
                                 current_user=user,
                                 companies=companies,
                                 departments=departments,
                                 financial_years=financial_years)

@app.route('/department-management')
@admin_required
def department_management():
    user = User.query.get(session['user_id'])
    departments = Department.query.all()
    
    return render_template('department_management.html',
                                 current_user=user,
                                 departments=departments)

@app.route('/add-company', methods=['POST'])
@admin_required
def add_company():
    company_name = request.form.get('company_name')
    
    if not company_name:
        flash('Company name is required.', 'error')
    elif Company.query.filter_by(CompanyName=company_name).first():
        flash('Company already exists.', 'error')
    else:
        company = Company(CompanyName=company_name, ActiveFlag=True)  # type: ignore
        db.session.add(company)
        db.session.commit()
        flash('Company added successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/toggle-company/<int:company_id>', methods=['POST'])
@admin_required
def toggle_company(company_id):
    company = Company.query.get_or_404(company_id)
    company.ActiveFlag = not company.ActiveFlag
    db.session.commit()
    flash(f'Company {company.CompanyName} {"activated" if company.ActiveFlag else "deactivated"}.', 'success')
    return redirect(url_for('config_master'))

@app.route('/edit-company/<int:company_id>', methods=['POST'])
@admin_required
def edit_company(company_id):
    company = Company.query.get_or_404(company_id)
    new_name = request.form.get('company_name')
    
    if not new_name:
        flash('Company name is required.', 'error')
    elif new_name != company.CompanyName and Company.query.filter_by(CompanyName=new_name).first():
        flash('Company name already exists.', 'error')
    else:
        company.CompanyName = new_name
        db.session.commit()
        flash(f'Company updated successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/delete-company/<int:company_id>', methods=['POST'])
@admin_required
def delete_company(company_id):
    company = Company.query.get_or_404(company_id)
    
    db.session.delete(company)
    db.session.commit()
    flash(f'Company {company.CompanyName} deleted successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/add-department', methods=['POST'])
@admin_required
def add_department():
    dept_name = request.form.get('dept_name')
    
    if not dept_name:
        flash('Department name is required.', 'error')
    elif Department.query.filter_by(DeptName=dept_name).first():
        flash('Department already exists.', 'error')
    else:
        department = Department(DeptName=dept_name, ActiveFlag=True)  # type: ignore
        db.session.add(department)
        db.session.commit()
        flash('Department added successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/toggle-department/<int:dept_id>', methods=['POST'])
@admin_required
def toggle_department(dept_id):
    dept = Department.query.get_or_404(dept_id)
    dept.ActiveFlag = not dept.ActiveFlag
    db.session.commit()
    flash(f'Department {dept.DeptName} {"activated" if dept.ActiveFlag else "deactivated"}.', 'success')
    return redirect(url_for('config_master'))

@app.route('/edit-department/<int:dept_id>', methods=['POST'])
@admin_required
def edit_department(dept_id):
    dept = Department.query.get_or_404(dept_id)
    new_name = request.form.get('dept_name')
    
    if not new_name:
        flash('Department name is required.', 'error')
    elif new_name != dept.DeptName and Department.query.filter_by(DeptName=new_name).first():
        flash('Department name already exists.', 'error')
    else:
        dept.DeptName = new_name
        db.session.commit()
        flash(f'Department updated successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/delete-department/<int:dept_id>', methods=['POST'])
@admin_required
def delete_department(dept_id):
    dept = Department.query.get_or_404(dept_id)
    
    # Check if department has associated users or uploads
    if dept.users or dept.uploads or dept.templates:
        flash('Cannot delete department with associated users, uploads, or templates. Deactivate it instead.', 'error')
    else:
        db.session.delete(dept)
        db.session.commit()
        flash(f'Department {dept.DeptName} deleted successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/add-fy', methods=['POST'])
@admin_required
def add_fy():
    fy_name = request.form.get('fy_name')
    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')
    
    if not fy_name or not start_date_str or not end_date_str:
        flash('All fields are required.', 'error')
        return redirect(url_for('config_master'))
    
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    if FinancialYear.query.filter_by(FYName=fy_name).first():
        flash('Financial Year already exists.', 'error')
    elif start_date >= end_date:
        flash('Start date must be before end date.', 'error')
    else:
        fy = FinancialYear(FYName=fy_name, StartDate=start_date, EndDate=end_date, ActiveFlag=False)  # type: ignore
        db.session.add(fy)
        db.session.commit()
        flash('Financial Year added successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/toggle-fy/<int:fy_id>', methods=['POST'])
@admin_required
def toggle_fy(fy_id):
    fy = FinancialYear.query.get_or_404(fy_id)
    
    if not fy.ActiveFlag:
        FinancialYear.query.update({FinancialYear.ActiveFlag: False})
        fy.ActiveFlag = True
        db.session.commit()
        flash(f'Financial Year {fy.FYName} is now active.', 'success')
    else:
        flash('This Financial Year is already active.', 'error')
    
    return redirect(url_for('config_master'))

@app.route('/edit-fy/<int:fy_id>', methods=['POST'])
@admin_required
def edit_fy(fy_id):
    fy = FinancialYear.query.get_or_404(fy_id)
    fy_name = request.form.get('fy_name')
    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')
    
    if not fy_name or not start_date_str or not end_date_str:
        flash('All fields are required.', 'error')
        return redirect(url_for('config_master'))
    
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    if fy_name != fy.FYName and FinancialYear.query.filter_by(FYName=fy_name).first():
        flash('Financial Year name already exists.', 'error')
    elif start_date >= end_date:
        flash('Start date must be before end date.', 'error')
    else:
        fy.FYName = fy_name
        fy.StartDate = start_date
        fy.EndDate = end_date
        db.session.commit()
        flash('Financial Year updated successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/delete-fy/<int:fy_id>', methods=['POST'])
@admin_required
def delete_fy(fy_id):
    fy = FinancialYear.query.get_or_404(fy_id)
    
    if fy.uploads:
        flash('Cannot delete Financial Year with associated uploads. Deactivate it instead.', 'error')
    else:
        db.session.delete(fy)
        db.session.commit()
        flash(f'Financial Year {fy.FYName} deleted successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/user-management')
@admin_required
def user_management():
    user = User.query.get(session['user_id'])
    users = User.query.all()
    departments = Department.query.all()
    roles = Role.query.all()
    
    return render_template('user_management.html',
                                 current_user=user,
                                 users=users,
                                 departments=departments,
                                 roles=roles)

@app.route('/add-user', methods=['POST'])
@admin_required
def add_user():
    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password')
    department_id = request.form.get('department_id')
    role_id = request.form.get('role_id')
    
    if not username or not email or not password or not department_id or not role_id:
        flash('All fields are required.', 'error')
    elif User.query.filter_by(Username=username).first():
        flash('Username already exists.', 'error')
    elif User.query.filter_by(Email=email).first():
        flash('Email already exists.', 'error')
    else:
        hashed_password = hash_password(password)
        new_user = User(  # type: ignore
            Username=username,
            Email=email,
            PasswordHash=hashed_password,
            DepartmentID=department_id,
            RoleID=role_id,
            IsActive=True
        )
        db.session.add(new_user)
        db.session.commit()
        flash('User created successfully!', 'success')
    
    return redirect(url_for('user_management'))

@app.route('/edit-user/<int:user_id>', methods=['POST'])
@admin_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    username = request.form.get('username')
    email = request.form.get('email')
    department_id = request.form.get('department_id')
    role_id = request.form.get('role_id')
    password = request.form.get('password')
    
    if not username or not email or not department_id or not role_id:
        flash('All fields are required.', 'error')
        return redirect(url_for('user_management'))
    
    if username != user.Username and User.query.filter_by(Username=username).first():
        flash('Username already exists.', 'error')
    elif email != user.Email and User.query.filter_by(Email=email).first():
        flash('Email already exists.', 'error')
    else:
        user.Username = username
        user.Email = email
        user.DepartmentID = department_id
        user.RoleID = role_id
        if password:
            user.PasswordHash = hash_password(password)
        db.session.commit()
        flash('User updated successfully!', 'success')
    
    return redirect(url_for('user_management'))

@app.route('/toggle-user/<int:user_id>', methods=['POST'])
@admin_required
def toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    user.IsActive = not user.IsActive
    db.session.commit()
    flash(f'User {user.Username} {"activated" if user.IsActive else "deactivated"}.', 'success')
    return redirect(url_for('user_management'))

@app.route('/delete-user/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    
    if user.uploads:
        flash('Cannot delete user with associated uploads. Deactivate instead.', 'error')
    else:
        db.session.delete(user)
        db.session.commit()
        flash(f'User {user.Username} deleted successfully!', 'success')
    
    return redirect(url_for('user_management'))

@app.route('/mis-upload')
@login_required
def mis_upload():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    upload_allowed, upload_message = check_upload_window()
    
    if user.role.RoleName == 'Admin':
        departments = Department.query.all()
        uploads = MISUpload.query.order_by(MISUpload.UploadDate.desc()).all()
    else:
        # HOD and User can only see their department
        departments = Department.query.filter_by(DeptID=user.DepartmentID).all()
        uploads = MISUpload.query.filter_by(DepartmentID=user.DepartmentID).order_by(MISUpload.UploadDate.desc()).all()
    
    financial_years = FinancialYear.query.all()
    
    return render_template('mis_upload.html',
                                 current_user=user,
                                 upload_allowed=upload_allowed,
                                 upload_message=upload_message,
                                 current_date=date.today().strftime('%Y-%m-%d'),
                                 departments=departments,
                                 financial_years=financial_years,
                                 uploads=uploads)

@app.route('/upload-mis', methods=['POST'])
@login_required
def upload_mis():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    upload_allowed, upload_message = check_upload_window()
    
    if not upload_allowed:
        flash(upload_message, 'error')
        return redirect(url_for('mis_upload'))
    
    if 'file' not in request.files:
        flash('No file selected.', 'error')
        return redirect(url_for('mis_upload'))
    
    file = request.files['file']
    
    if not file or not file.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('mis_upload'))
    
    if not (file.filename.endswith('.xls') or file.filename.endswith('.xlsx')):
        flash('Only .xls or .xlsx files are allowed.', 'error')
        return redirect(url_for('mis_upload'))
    
    department_id = request.form.get('department_id')
    month_id = request.form.get('month_id')
    fy_id = request.form.get('fy_id')
    
    if not department_id or not month_id or not fy_id:
        flash('All fields are required.', 'error')
        return redirect(url_for('mis_upload'))
    
    # HOD and User can only upload for their own department
    if user.role.RoleName in ['HOD', 'User'] and int(department_id) != user.DepartmentID:
        flash('You can only upload for your own department.', 'error')
        return redirect(url_for('mis_upload'))
    
    filename = secure_filename(f"{department_id}_{month_id}_{fy_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    
    # Validate Excel file content
    is_valid, validation_message = validate_excel_file(filepath)
    
    if not is_valid:
        os.remove(filepath)
        flash(f'Validation Error: {validation_message}', 'error')
        return redirect(url_for('mis_upload'))
    
    upload = MISUpload(  # type: ignore
        DepartmentID=department_id,
        MonthID=month_id,
        FYID=fy_id,
        UploadedBy=session['user_id'],
        FilePath=filepath,
        Status='Validated'
    )
    db.session.add(upload)
    db.session.commit()
    
    flash(f'✓ Validation Success: {validation_message} File uploaded successfully!', 'success')
    return redirect(url_for('mis_upload'))

@app.route('/template-management')
@admin_required
def template_management():
    user = User.query.get(session['user_id'])
    departments = Department.query.all()
    templates = Template.query.all()
    
    return render_template('template_management.html',
                                 current_user=user,
                                 departments=departments,
                                 templates=templates)

@app.route('/download-template/<int:dept_id>')
@login_required
def download_template(dept_id):
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    template = Template.query.filter_by(DepartmentID=dept_id).order_by(Template.UploadDate.desc()).first()
    
    if not template:
        flash('No template found for this department.', 'error')
        return redirect(url_for('mis_upload'))
    
    from flask import send_file
    try:
        return send_file(template.FilePath, as_attachment=True, download_name=f"MIS_Template_{dept_id}.xlsx")
    except Exception as e:
        flash(f'Error downloading template: {str(e)}', 'error')
        return redirect(url_for('mis_upload'))

@app.route('/upload-template', methods=['POST'])
@admin_required
def upload_template():
    if 'file' not in request.files:
        flash('No file selected.', 'error')
        return redirect(url_for('template_management'))
    
    file = request.files['file']
    
    if not file or not file.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('template_management'))
    
    if not (file.filename.endswith('.xls') or file.filename.endswith('.xlsx')):
        flash('Only .xls or .xlsx files are allowed.', 'error')
        return redirect(url_for('template_management'))
    
    department_id = request.form.get('department_id')
    
    if not department_id:
        flash('Department is required.', 'error')
        return redirect(url_for('template_management'))
    
    filename = secure_filename(f"template_{department_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    
    template = Template(  # type: ignore
        DepartmentID=department_id,
        FilePath=filepath
    )
    db.session.add(template)
    db.session.commit()
    
    flash('Template uploaded successfully!', 'success')
    return redirect(url_for('template_management'))

def init_db():
    with app.app_context():
        db.create_all()
        
        if Role.query.count() == 0:
            roles = [
                Role(RoleName='Admin'),  # type: ignore
                Role(RoleName='HOD'),  # type: ignore
                Role(RoleName='User')  # type: ignore
            ]
            db.session.add_all(roles)
            db.session.commit()
            print("Roles seeded.")
        
        if Department.query.count() == 0:
            departments = [
                Department(DeptName='Finance', ActiveFlag=True),  # type: ignore
                Department(DeptName='HR', ActiveFlag=True),  # type: ignore
                Department(DeptName='IT', ActiveFlag=True)  # type: ignore
            ]
            db.session.add_all(departments)
            db.session.commit()
            print("Departments seeded.")
        
        if Company.query.count() == 0:
            company = Company(CompanyName='Default Company', ActiveFlag=True)  # type: ignore
            db.session.add(company)
            db.session.commit()
            print("Company seeded.")
        
        if FinancialYear.query.count() == 0:
            fy = FinancialYear(  # type: ignore
                FYName='2024-2025',
                StartDate=date(2024, 4, 1),
                EndDate=date(2025, 3, 31),
                ActiveFlag=True
            )
            db.session.add(fy)
            db.session.commit()
            print("Financial Year seeded.")
        
        if User.query.count() == 0:
            admin_role = Role.query.filter_by(RoleName='Admin').first()
            hod_role = Role.query.filter_by(RoleName='HOD').first()
            user_role = Role.query.filter_by(RoleName='User').first()
            
            finance_dept = Department.query.filter_by(DeptName='Finance').first()
            hr_dept = Department.query.filter_by(DeptName='HR').first()
            it_dept = Department.query.filter_by(DeptName='IT').first()
            
            if admin_role and hod_role and user_role and finance_dept and hr_dept and it_dept:
                users = [
                    User(  # type: ignore
                        Username='admin',
                        Email='admin@example.com',
                        PasswordHash=hash_password('admin123'),
                        DepartmentID=it_dept.DeptID,
                        RoleID=admin_role.RoleID,
                        IsActive=True
                    ),
                    User(  # type: ignore
                        Username='hod',
                        Email='hod@example.com',
                        PasswordHash=hash_password('hod123'),
                        DepartmentID=finance_dept.DeptID,
                        RoleID=hod_role.RoleID,
                        IsActive=True
                    ),
                    User(  # type: ignore
                        Username='user',
                        Email='user@example.com',
                        PasswordHash=hash_password('user123'),
                        DepartmentID=hr_dept.DeptID,
                        RoleID=user_role.RoleID,
                        IsActive=True
                    )
                ]
                db.session.add_all(users)
                db.session.commit()
                print("✓ Test users created:")
                print("  - Admin: username='admin', password='admin123'")
                print("  - HOD: username='hod', password='hod123'")
                print("  - User: username='user', password='user123'")
        
        print("Database initialized successfully!")

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
